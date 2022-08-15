import streamlit as st
from ics import Calendar, Event 
import pandas as pd
import numpy as np
from io import BytesIO
import openpyxl as op
from datetime import datetime
from tempfile import NamedTemporaryFile
from fpdf import FPDF
from fpdf.enums import XPos, YPos


def get_employer_data(excel, month: int, year: int):
    year = year if month else year-1
    month = month if month else 12
    
    sheet1 = pd.read_excel(excel, usecols='A:B', index_col = 0, header = None)
    try:
        sheet2 = pd.read_excel(excel, usecols='A:Y', sheet_name = f'Leave_Record_{year}', skiprows=np.arange(0), header=1)
    except ValueError:
        st.warning(f'Worksheet named Leave_Record_{year} not found')
        st.stop()

    # get name and start time 
    name = sheet1.loc['Employee'][1]
    first_day = sheet1.loc['Start'][1]

    # get sick_leave_balance and annaul_balance from last month
    sick_balance = sheet2['Sick' if month == 1 else f'Sick.{month-1}'][33]
    annaul_balance = sheet2[f'Vacation' if month == 1 else f'Vacation.{month-1}'][33]

    return name, first_day, sick_balance, annaul_balance

# get leave this month 
def calacuate_leave(employer_events: list) -> tuple[list, list]:
    sick_leave = []
    vacacies = []
    for event in employer_events:
        diff = event.end - event.begin
        diff = diff.seconds/3600 + diff.days*24
        date = event.begin

        while diff > 0:
            time = diff 
            if diff < 3:
                break
            time = 1 if diff > 6 else 0.5
            if 'sick leave' in event.name:         
                sick_leave.append((date, time))
            else:
                vacacies.append((date, time))
            date = date.shift(days = 1)
            diff -= 24 

    return sick_leave, vacacies

def update_excel(
        wb, month:int, year:int, sick_balance:float,
        annual_balance:float, sick_leave:list, vacaies:list, 
        name:str, worked_month:int , worked_year:int, first_day):

    # calculate the col of this month in sheet Leave_Record_[Year]
    vacaies_col= chr(2*(month-1)+ord('B'))
    sick_col = chr(ord(vacaies_col)+1)

    month_list = ['Jan','Feb','Mar',
            'Apr', 'May', 'Jun',
            'Jul' , 'Aug', 'Sep',
            'Oct', 'Nov', 'Dec']
    # create a new sheet in jan
    if f'Leave_Record_{year}' not in wb.sheetnames:
        wb.create_sheet(f'Leave_Record_{year}')
        for i, m in enumerate(month_list):
            new_col = chr(2*(i)+ord('B'))
            wb[f'Leave_Record_{year}'][f'{new_col}1'] = m
            wb[f'Leave_Record_{year}'][f'{new_col}2'] = 'Vacation' 
            wb[f'Leave_Record_{year}'][f'{chr(ord(new_col)+1)}2'] = 'Sick' 
        for i in range(31):
            wb[f'Leave_Record_{year}'][f'A{i+3}'] = i+1 
        wb[f'Leave_Record_{year}']['A34'] = 'Earned'
        wb[f'Leave_Record_{year}']['A35'] = 'Used'
        wb[f'Leave_Record_{year}']['A36'] = 'Balance'

    # update vacation on sheet Leave_Record_[Year] 
    vacacies_count = 0
    for start, time in vacaies:
        wb[f'Leave_Record_{year}'][f'{vacaies_col}{start.day+2}'] = time
        vacacies_count += time
    earned_vacacies = 0
    if worked_month == 0:
        earned_vacacies = 7 if year < 2 else 7 + worked_year - 1 
        earned_vacacies = earned_vacacies if earned_vacacies <= 14 else 14
    wb[f'Leave_Record_{year}'][f'{vacaies_col}{34}'] = earned_vacacies 
    wb[f'Leave_Record_{year}'][f'{vacaies_col}{35}'] = vacacies_count 
    updated_vacacies =  annual_balance + earned_vacacies - vacacies_count 
    updated_vacacies = 14 if updated_vacacies > 14 else updated_vacacies
    wb[f'Leave_Record_{year}'][f'{vacaies_col}{36}'] =  updated_vacacies

    # update vacaction on first sheet
    col = chr(worked_year+ord('C'))
    if worked_month == 0: 
        wb[name][f'{col}12'] = earned_vacacies 
    try:
        wb[name][f'{col}13'] = wb[name][f'{col}13'].value + vacacies_count
    except TypeError:
        wb[name][f'{col}13'] = vacacies_count
    wb[name][f'{col}15'] = updated_vacacies 

    # update sick leave on sheet Leave_Record_[Year]
    sick_count = 0
    for start, time in sick_leave:
        wb[f'Leave_Record_{year}'][f'{sick_col}{start.day+2}'] = time
        sick_count += time
    earned_sick_leave = 4 if worked_year >= 1 else 2
    updated_sick_leave = sick_balance + earned_sick_leave - sick_count 
    if  updated_sick_leave >= 120:
        earned_sick_leave = earned_sick_leave - (updated_sick_leave - 120) 
    updated_sick_leave = 120 if updated_sick_leave > 120 else updated_sick_leave
    wb[f'Leave_Record_{year}'][f'{sick_col}{34}'] = earned_sick_leave 
    wb[f'Leave_Record_{year}'][f'{sick_col}{35}'] = sick_count
    wb[f'Leave_Record_{year}'][f'{sick_col}{36}'] = updated_sick_leave 

    #update sick leave on first sheet
    col = chr(year-first_day.year+ord('C'))
    row = month + 20
    wb[name][f'{col}{row}'] =earned_sick_leave 
    try:
        wb[name][f'{col}{34}'] = wb[name][f'{col}{34}'].value + sick_count 
    except TypeError:
        wb[name][f'{col}{34}'] = sick_count 
    wb[name]['E35'] = updated_sick_leave 

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        data = BytesIO(tmp.read())
    st.download_button("Retrieve excel file",
        data=data,
        mime='xlsx',
        file_name=f"Vacation_Sick_Record_{name}_{month_list[month-1]}-{year}.xlsx")

    return earned_vacacies, earned_sick_leave, updated_vacacies, updated_sick_leave, vacacies_count, sick_count

def generate_pdf(
        month:str, year:int, updated_vacacies:float,
        updated_sick_leave:float, sick_leave:list, vacaies:list, 
        name:str, earned_vacacies:float,  earned_sick_leave:float,
        vacaies_count:float, sick_count:float):

    file_name = f'Leave_Report_{name}_{month}{year}'

    # create pdf
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(0)
    # header
    pdf.image('title.png', 7, 5, 200)
    pdf.set_font('helvetica', 'B', 18)

    # report info
    info = f'Leave Report\n{month} {year}\nEmployee: {name}'
    pdf.cell(0, 15, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.multi_cell(100, 10, info)
    pdf.cell(0, 5, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font('helvetica', size=12)

    # store table cell data in a list
    data = [[' ', 'Vacation', 'Sick']]

    vacaies_dict = {start.day: time for start, time in vacaies}
    sick_dict = {start.day: time for start, time in sick_leave}
      
    for i in range(1, 32):
        row = [str(i)]
        row.append(str(vacaies_dict[i]) if i in vacaies_dict else ' ')
        row.append(str(sick_dict[i]) if i in sick_dict else ' ')
        data.append(row)        
    data.append(['Used this Month (Day)', str(vacacies_count), str(sick_count)])
    data.append(['Earned this Month (Day)', str(earned_vacacies), str(earned_sick_leave)])
    data.append(['Up to day Balance (Day)', str(updated_vacacies), str(updated_sick_leave)])

    # generate table 
    line_height = pdf.font_size * 1.3 
    col_width = pdf.epw / 3  
    for i, row in enumerate(data):
        for j, datum in enumerate(row):
            if i == 0 or (j==0 and i > 31):
                pdf.multi_cell(col_width, line_height, datum, border=1,
                        new_x="RIGHT", new_y="TOP", max_line_height=pdf.font_size)
            else:
                pdf.multi_cell(col_width, line_height, datum, border=1,
                        new_x="RIGHT", new_y="TOP", max_line_height=pdf.font_size, align='R')
        pdf.ln(line_height)

    # footer
    pdf.set_line_width(0.25)
    pdf.set_draw_color(r=20, g=200, b=200)
    pdf.line(x1=10,y1=270, x2=200, y2=270)
    pdf.cell(0, 18, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font('helvetica', size=10.5)
    footer_text = 'Room 627, 6/F, 17 Science Park West Avenue, Hong Kong Science Park, Sha Tin, Hong Kong\ninfo@decodecure.com +852 3703 2570'
    pdf.multi_cell(pdf.epw, 5, footer_text, align='C')

    # output pdf
    output = BytesIO()
    pdf.output(output)
    st.download_button("Retrieve PDF",
        data=output.getvalue(),
        mime='pdf',
        file_name=f"{file_name}.pdf")

  
# App Start ######################
st.title("Employee Leave Report")
st.markdown("""---""")

# input data 
col1, col2 = st.columns([1,1])
with col1:
    calendar = st.file_uploader('ics')
with col2:
    excel = st.file_uploader('excel')

if not calendar or not excel:
    st.stop()

wb = op.load_workbook(filename=BytesIO(excel.read()))

month_list = {'Jan': 1,'Feb': 2,'Mar': 3,
    'Apr': 4, 'May': 5, 'Jun': 6,
    'Jul': 7, 'Aug': 8, 'Sep': 9,
    'Oct': 10, 'Nov': 11, 'Dec': 12}

# check current month from excel file name
month = 1 
for m in month_list:
    if m in excel.name:
        month = month_list[m] + 1
month = month if month <= 12 else 1  

# allow user to input custom year and month
month = st.selectbox('Month', month_list.keys(), index=month-1) 
month_name = month
month = month_list[month]
year = datetime.now().year
year = int(st.text_input('year', value=year))

# get employer data from record
name, first_day, sick_balance, annual_balance = get_employer_data(excel, month-1, year)

# calculate the employer worked time
worked_year = year - first_day.year 
worked_month = month - first_day.month
if worked_month < 0:
    worked_year -= 1 
    worked_month = 12 - first_day.month + month 
#  get event in caledar
calendar = Calendar(calendar.read().decode('utf-8'))
events = calendar.events

# get all leave of employer this month 
employer_events = []
for event in events:
    if (month == event.begin.month and year == event.begin.year and name in event.name):
        employer_events.append(event)

sick_leave, vacacies = calacuate_leave(employer_events)
earned_vacacies, earned_sick_leave, updated_vacacies, updated_sick_leave, vacacies_count, sick_count = update_excel(wb, month, year,
        sick_balance, annual_balance, sick_leave,
        vacacies, name, worked_month, worked_year, first_day)
generate_pdf(month_name, year, updated_vacacies,
        updated_sick_leave, sick_leave, vacacies, 
        name, earned_vacacies,  earned_sick_leave,
        vacacies_count, sick_count)