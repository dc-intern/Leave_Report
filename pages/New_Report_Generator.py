from openpyxl import Workbook 
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.styles import NamedStyle
import streamlit as st
from tempfile import NamedTemporaryFile
from datetime import datetime
from io import BytesIO

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
                     
month_list = ['Jan','Feb','Mar',
        'Apr', 'May', 'Jun',
        'Jul' , 'Aug', 'Sep',
        'Oct', 'Nov', 'Dec']

st.title("New Report Generator")
st.markdown("""---""")

name = st.text_input('Employee Name', value='Name')
nickname = st.text_input('Employee Nickname', value='Nickname')
first_day = st.date_input('First Day of Employee')

wbook = Workbook()
ws = wbook.worksheets[0]
ws.title = name 

ws['A1'] = 'Employee'
ws['B1'] = name
ws['C1'] = nickname
ws['A2'] = 'Start'
ws['B2'] = first_day
ws['A3'] = 'Annual Leave (Cap, day)'
ws['B3'] = '14'
ws['A4'] = 'Sick Leave (Cap, day)'
ws['B4'] = '120'

# generate vacacies table
date_style = NamedStyle(name='custom_datetime', number_format='MM/YYYY')
for i in range(9):
    col = chr(ord('B')+i)
    year = first_day.replace(year=first_day.year+i)
    ws[f'{col}8'] = year
    ws[f'{col}8'].style = date_style
    ws[f'{col}9'] = i
data = ['Anniversary', 'Leave entitlement begin', 'Current Year entitlement', 
        'Annual Leave (Day)', 'Leave Days', 'Remark',
        'Balance']
for i, datum in enumerate(data):
    ws[f'A{i+9}'] = datum
    ws[f'A{i+9}'].border = thin_border 

# apply boarder
for col in 'BCDEFGHIJ':
    for row in range(8,16):
        ws[f'{col}{row}'].border = thin_border

# generate sick leave table
ws['A19'] = 'Sick Leave (earned, day)'
ws['A19'].border = thin_border
ws['D35'] = 'Balance'
ws['D35'].border = thin_border
ws['E35'].border = thin_border
ws['B33'] = 'Total'
ws['B33'].border = thin_border
ws['B34'] = 'Leave'
ws['B34'].border = thin_border
for i in range(8):
    col = chr(ord('C')+i)
    ws[f'{col}20'] = first_day.year+i
for i, month in enumerate(month_list):
    ws[f'B{i+21}'] = month
for col in 'CDEFGHIJ':
    ws[f'{col}33'] = f'=sum({col}21:{col}32)'

# apply boarder
for col in 'BCDEFGHIJ':
    for row in range(19,35):
        ws[f'{col}{row}'].border = thin_border

# fix col width
ws.column_dimensions['A'].width = 26
for col in 'BCDEFGHIJKLMNOPQRSTUVWXYZ':
    ws.column_dimensions[col].width = 13

# fix font type
ft = Font(name='Calibri')
for col in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
    for row in range(36):
        ws[f'{col}{row+1}'].font = ft

# generate downloard button   
with NamedTemporaryFile() as tmp:
    wbook.save(tmp.name)
    data = BytesIO(tmp.read())
st.download_button("Retrieve excel file",
    data=data,
    mime='xlsx',
    file_name=f"Vacation_Sick_Record_{name}_Empty.xlsx")
